"""
Import benchmark and observed values from FSFSI_indicator_level_parameters.xlsx.

Reads only the Indicator_Parameters sheet. Updates:
- Indicator: unit, higher_is_better (from Direction), default_sensitivity (from alpha_per_bnLCU)
- IndicatorData: observed_value (Obs_value), benchmark_value (Benchmark_used), sensitivity_parameter;
  creates IndicatorData for (indicator, Obs_year) if missing, using funding/records from the sheet.

Usage:
    python manage.py import_indicator_parameters /path/to/FSFSI_indicator_level_parameters.xlsx
"""
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.db.models import Q
from django.utils import timezone

from apps.fsfvi_data.models import Indicator, IndicatorData, IndicatorComponent, DataStatus


# Column indices (0-based) for Indicator_Parameters sheet
COL_INDICATOR_CODE = 0   # A
COL_INDICATOR_NAME = 1   # B
COL_COMPONENT = 2        # C
COL_FUNDING_WEIGHTED_BN = 3   # D
COL_FUNDING_GROSS_BN = 4     # E
COL_RECORDS = 5          # F
COL_FALLBACK_RECORDS = 6 # G
COL_FSCI_INDICATOR_USED = 7  # H
COL_OBS_VALUE = 8        # I
COL_OBS_YEAR = 9         # J
COL_OBS_UNIT = 10        # K
COL_BENCHMARK_USED = 11  # L (numeric benchmark value)
COL_BENCHMARK_TYPE = 12  # M
COL_DIRECTION = 13       # N  "higher" | "lower"
COL_DELTA = 14           # O
COL_DELTA_IMPUTED = 15   # P
COL_ALPHA_PER_BN_LCU = 16   # Q → default_sensitivity / sensitivity_parameter


def _safe_decimal(value, default=None):
    if value is None:
        return default
    if isinstance(value, Decimal):
        return value
    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError):
        return default


def _safe_int(value, default=0):
    if value is None:
        return default
    try:
        return int(round(float(value)))
    except (TypeError, ValueError):
        return default


def _safe_str(value, max_len=None):
    s = "" if value is None else str(value).strip()
    if max_len and len(s) > max_len:
        return s[:max_len]
    return s


def _direction_to_higher_is_better(direction):
    if not direction:
        return True
    return str(direction).strip().lower() == "higher"


def _component_label_to_value(label):
    if not label:
        return ""
    label = str(label).strip()
    for choice in IndicatorComponent:
        if choice.label == label:
            return choice.value
    normalized = label.lower().replace(" ", "_").replace("-", "_")
    if normalized in [c.value for c in IndicatorComponent]:
        return normalized
    return label.lower().replace(" ", "_").replace("-", "_")[:30]


class Command(BaseCommand):
    help = "Import benchmark and observed values from FSFSI_indicator_level_parameters.xlsx (Indicator_Parameters sheet)"

    def add_arguments(self, parser):
        parser.add_argument(
            "excel_path",
            type=str,
            help="Path to FSFSI_indicator_level_parameters.xlsx",
        )
        parser.add_argument(
            "--default-fiscal-year",
            type=int,
            default=None,
            help="If Obs_year is blank in the sheet, use this year for IndicatorData (e.g. 2024).",
        )

    def handle(self, *args, **options):
        path = Path(options["excel_path"])
        if not path.is_file():
            raise CommandError(f"File not found: {path}")

        try:
            import openpyxl
        except ImportError:
            raise CommandError("openpyxl is required. Install with: pip install openpyxl")

        self.stdout.write(f"Loading workbook (read_only): {path}")
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

        if "Indicator_Parameters" not in wb.sheetnames:
            wb.close()
            raise CommandError("Sheet 'Indicator_Parameters' not found")

        ws = wb["Indicator_Parameters"]
        default_fy = options.get("default_fiscal_year")
        try:
            with transaction.atomic():
                self._load_indicator_parameters(ws, default_fiscal_year=default_fy)
        finally:
            wb.close()

        self.stdout.write(self.style.SUCCESS("Import completed successfully."))

    def _load_indicator_parameters(self, ws, default_fiscal_year=None):
        """Read Indicator_Parameters rows and update Indicator + IndicatorData."""
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) <= COL_BENCHMARK_USED:
                continue
            code = _safe_str(row[COL_INDICATOR_CODE], 20)
            if not code or not code.upper().startswith("IND-"):
                continue

            obs_year = _safe_int(row[COL_OBS_YEAR])
            if not obs_year or obs_year < 2000 or obs_year > 2100:
                obs_year = default_fiscal_year

            obs_value = _safe_decimal(row[COL_OBS_VALUE])
            benchmark_raw = row[COL_BENCHMARK_USED] if len(row) > COL_BENCHMARK_USED else None
            # Benchmark_used can be numeric or string (e.g. "SSA_10/90pct"); store numeric for benchmark_value
            benchmark_value = _safe_decimal(benchmark_raw)
            # When benchmark_value is empty, Benchmark_used_type (e.g. Global_10/90pct) tells us the reference
            benchmark_used_type = _safe_str(row[COL_BENCHMARK_TYPE], 100) if len(row) > COL_BENCHMARK_TYPE else ""
            direction = _safe_str(row[COL_DIRECTION]) if len(row) > COL_DIRECTION else ""
            obs_unit = _safe_str(row[COL_OBS_UNIT], 50) if len(row) > COL_OBS_UNIT else ""
            alpha = _safe_decimal(row[COL_ALPHA_PER_BN_LCU]) if len(row) > COL_ALPHA_PER_BN_LCU else None

            funding_weighted = _safe_decimal(row[COL_FUNDING_WEIGHTED_BN], Decimal("0")) if len(row) > COL_FUNDING_WEIGHTED_BN else Decimal("0")
            funding_gross = _safe_decimal(row[COL_FUNDING_GROSS_BN], Decimal("0")) if len(row) > COL_FUNDING_GROSS_BN else Decimal("0")
            records = _safe_int(row[COL_RECORDS]) if len(row) > COL_RECORDS else 0
            fallback = _safe_int(row[COL_FALLBACK_RECORDS]) if len(row) > COL_FALLBACK_RECORDS else 0
            component_label = _safe_str(row[COL_COMPONENT]) if len(row) > COL_COMPONENT else ""
            component_value = _component_label_to_value(component_label)

            rows.append({
                "code": code,
                "obs_year": obs_year,
                "observed_value": obs_value,
                "benchmark_value": benchmark_value,
                "benchmark_used_type": benchmark_used_type,
                "higher_is_better": _direction_to_higher_is_better(direction),
                "unit": obs_unit,
                "default_sensitivity": alpha,
                "funding_weighted_bn": funding_weighted,
                "funding_gross_bn": funding_gross,
                "records_count": records,
                "fallback_records": fallback,
                "component": component_value,
            })

        if not rows:
            self.stdout.write(self.style.WARNING("No data rows found in Indicator_Parameters."))
            return

        codes = list({r["code"] for r in rows})
        indicators = {ind.code: ind for ind in Indicator.objects.filter(code__in=codes).only("id", "code", "unit", "higher_is_better", "default_sensitivity")}
        if not indicators:
            self.stdout.write(self.style.WARNING("No matching indicators in DB. Run import_budget_mapping first to create indicators."))
            return

        # Build (indicator_id, fiscal_year) -> row for updates/creates
        by_key = {}
        for r in rows:
            if r["code"] not in indicators:
                continue
            key = (indicators[r["code"]].id, r["obs_year"] or 0)
            if key not in by_key:
                by_key[key] = r
            else:
                # Keep row with obs_year set; prefer one with observed_value/benchmark
                existing = by_key[key]
                if (r["observed_value"] is not None or r["benchmark_value"] is not None) and (existing["observed_value"] is None and existing["benchmark_value"] is None):
                    by_key[key] = r

        # Update Indicator: unit, higher_is_better, default_sensitivity (per code, last row wins per code)
        indicator_updates = {}
        for r in rows:
            if r["code"] not in indicators:
                continue
            ind = indicators[r["code"]]
            indicator_updates[ind.id] = {
                "unit": r["unit"] or ind.unit,
                "higher_is_better": r["higher_is_better"],
                "default_sensitivity": r["default_sensitivity"] if r["default_sensitivity"] is not None else ind.default_sensitivity,
            }

        for ind in indicators.values():
            if ind.id in indicator_updates:
                u = indicator_updates[ind.id]
                ind.unit = u["unit"]
                ind.higher_is_better = u["higher_is_better"]
                ind.default_sensitivity = u["default_sensitivity"]
        Indicator.objects.bulk_update(
            [indicators[c] for c in codes if c in indicators],
            ["unit", "higher_is_better", "default_sensitivity"],
        )
        self.stdout.write(f"  Indicators updated: {len(indicator_updates)}")

        # IndicatorData: existing (indicator_id, fiscal_year) -> update; else create
        existing_data = {}
        for ind_id, fy in by_key:
            if fy <= 0:
                continue
            existing_data[(ind_id, fy)] = None  # placeholder
        if existing_data:
            keys = list(existing_data.keys())
            q = Q()
            for (ind_id, fy) in keys:
                q |= Q(indicator_id=ind_id, fiscal_year=fy)
            for rec in IndicatorData.objects.filter(q).select_related("indicator"):
                existing_data[(rec.indicator_id, rec.fiscal_year)] = rec

        to_update = []
        to_create = []
        for (ind_id, fiscal_year), r in by_key.items():
            if fiscal_year <= 0:
                continue
            rec = existing_data.get((ind_id, fiscal_year))
            if rec:
                rec.observed_value = r["observed_value"]
                rec.benchmark_value = r["benchmark_value"]
                rec.benchmark_used_type = r.get("benchmark_used_type") or ""
                if r["default_sensitivity"] is not None:
                    rec.sensitivity_parameter = r["default_sensitivity"]
                rec.updated_at = timezone.now()
                to_update.append(rec)
            else:
                to_create.append(
                    IndicatorData(
                        indicator_id=ind_id,
                        fiscal_year=fiscal_year,
                        records_count=r["records_count"],
                        fallback_records=r["fallback_records"],
                        gross_lcu_bn=r["funding_gross_bn"],
                        weighted_lcu_bn=r["funding_weighted_bn"],
                        share_weighted_percent=Decimal("0"),
                        observed_value=r["observed_value"],
                        benchmark_value=r["benchmark_value"],
                        benchmark_used_type=r.get("benchmark_used_type") or "",
                        sensitivity_parameter=r["default_sensitivity"],
                        status=DataStatus.VALIDATED,
                    )
                )

        if to_update:
            IndicatorData.objects.bulk_update(
                to_update,
                ["observed_value", "benchmark_value", "benchmark_used_type", "sensitivity_parameter", "updated_at"],
                batch_size=500,
            )
            self.stdout.write(f"  IndicatorData updated: {len(to_update)}")
        if to_create:
            IndicatorData.objects.bulk_create(to_create)
            self.stdout.write(f"  IndicatorData created: {len(to_create)}")
        if not to_update and not to_create:
            self.stdout.write(self.style.WARNING("  No IndicatorData rows updated or created (fiscal years may not match existing data)."))
