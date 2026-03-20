"""
Import budget lines to food system indicators mapping from Excel into the database.

Optimized for speed: read_only workbook, bulk_create in batches, single transaction.
Sheets: Mapping (raw budget lines), Indicator_Summary (aggregated per indicator).

Usage:
    python manage.py import_budget_mapping /path/to/budget_lines_to_food_system_indicators_mapping.xlsx
    python manage.py import_budget_mapping /path/to/file.xlsx --fiscal-year 2018
    python manage.py import_budget_mapping /path/to/file.xlsx --summary-only  # Skip Mapping sheet
"""
import re
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from apps.fsfvi_data.models import (
    BudgetLineMapping,
    Indicator,
    IndicatorData,
    IndicatorComponent,
    DataStatus,
)


# Primary_indicator format: "IND-16: Share of production marketed (%)"
INDICATOR_CODE_RE = re.compile(r"^(IND-\d+)\s*:?\s*(.*)$", re.I)


def _parse_primary_indicator(value):
    """Return (code, name) e.g. ('IND-16', 'Share of production marketed (%)')."""
    if not value or not str(value).strip():
        return None, None
    s = str(value).strip()
    m = INDICATOR_CODE_RE.match(s)
    if m:
        return m.group(1).strip(), (m.group(2) or "").strip() or s
    if s.upper().startswith("IND-"):
        return s.split(":", 1)[0].strip(), s
    return None, s


def _component_label_to_value(label):
    """Map Excel component label to IndicatorComponent value."""
    if not label:
        return ""
    label = str(label).strip()
    for choice in IndicatorComponent:
        if choice.label == label:
            return choice.value
    # Normalize: "Crop Production" -> crop_production
    normalized = label.lower().replace(" ", "_").replace("-", "_")
    if normalized in [c.value for c in IndicatorComponent]:
        return normalized
    return label.lower().replace(" ", "_").replace("-", "_")[:30]


def _safe_decimal(value, default=None):
    if value is None:
        return default
    if isinstance(value, Decimal):
        return value
    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError):
        return default


def _safe_int(value, default=0):
    if value is None:
        return default
    try:
        return int(round(float(value)))
    except (TypeError, ValueError):
        return default


def _safe_str(value, max_len=None):
    s = "" if value is None else str(value).strip()
    if max_len and len(s) > max_len:
        return s[:max_len]
    return s


class Command(BaseCommand):
    help = "Import budget_lines_to_food_system_indicators_mapping.xlsx into Indicator, IndicatorData, and BudgetLineMapping"

    def add_arguments(self, parser):
        parser.add_argument(
            "excel_path",
            type=str,
            help="Path to budget_lines_to_food_system_indicators_mapping.xlsx",
        )
        parser.add_argument(
            "--fiscal-year",
            type=int,
            default=None,
            help="Import only this fiscal year. If omitted, imports ALL years found in the Mapping sheet.",
        )
        parser.add_argument(
            "--summary-only",
            action="store_true",
            help="Only load Indicator_Summary sheet (Indicator + IndicatorData). Skip Mapping sheet.",
        )
        parser.add_argument(
            "--batch-size",
            type=int,
            default=2000,
            help="bulk_create batch size (default 2000).",
        )

    def handle(self, *args, **options):
        path = Path(options["excel_path"])
        if not path.is_file():
            raise CommandError(f"File not found: {path}")

        try:
            import openpyxl
        except ImportError:
            raise CommandError(
                "openpyxl is required. Install with: pip install openpyxl"
            )

        fiscal_year = options["fiscal_year"]
        summary_only = options["summary_only"]
        batch_size = max(500, min(5000, options["batch_size"]))

        self.stdout.write(f"Loading workbook (read_only): {path}")
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

        try:
            with transaction.atomic():
                if "Indicator_Summary" not in wb.sheetnames:
                    raise CommandError("Sheet 'Indicator_Summary' not found")
                if not summary_only and "Mapping" not in wb.sheetnames:
                    raise CommandError("Sheet 'Mapping' not found")

                # Always load Indicator_Summary first to create/update Indicator records
                # Use fiscal_year if specified, otherwise a default for the summary
                summary_fy = fiscal_year or 2024
                self._load_indicator_summary(wb["Indicator_Summary"], summary_fy, batch_size)

                if not summary_only:
                    # Aggregate budget data per year from Mapping sheet
                    self._load_mapping_per_year(wb["Mapping"], fiscal_year, batch_size)
        finally:
            wb.close()

        self.stdout.write(self.style.SUCCESS("Import completed successfully."))

    def _infer_fiscal_year(self, ws):
        """Read first data row YEAR (column A) e.g. '2018/2019' -> 2018."""
        for row in ws.iter_rows(min_row=2, max_row=10, min_col=1, max_col=1, values_only=True):
            val = row[0] if row else None
            if not val:
                continue
            s = str(val).strip()
            m = re.match(r"^(\d{4})", s)
            if m:
                return int(m.group(1))
        return None

    def _load_indicator_summary(self, ws, fiscal_year, batch_size):
        """Load Indicator_Summary sheet into Indicator + IndicatorData. Replace strategy for that fiscal_year."""
        # Columns: A=Primary_indicator, B=Indicator_component, C=Indicator, D=Records, E=Fallback_records,
        #          F=Gross_LCU_bn, G=Weighted_LCU_bn, H=Share_weighted_%
        summary_rows = []
        seen_codes = set()
        row_num = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_num += 1
            if not row or len(row) < 8:
                continue
            primary_ind, comp_label, ind_name, records, fallback, gross_bn, weighted_bn, share_pct = (
                row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7]
            )
            code, parsed_name = _parse_primary_indicator(primary_ind)
            if not code:
                continue
            name = _safe_str(ind_name or parsed_name, 255) or parsed_name or code
            component_value = _component_label_to_value(comp_label)
            if not component_value:
                continue
            if code not in seen_codes:
                seen_codes.add(code)
                summary_rows.append({
                    "code": code,
                    "name": name,
                    "component": component_value,
                    "display_order": row_num,
                    "records": records,
                    "fallback": fallback,
                    "gross_bn": gross_bn,
                    "weighted_bn": weighted_bn,
                    "share_pct": share_pct,
                })

        if not summary_rows:
            self.stdout.write(self.style.WARNING("  No data rows in Indicator_Summary."))
            return

        # Ensure all indicators exist (bulk create missing ones)
        existing = {ind.code: ind.id for ind in Indicator.objects.only("id", "code").iterator(chunk_size=500)}
        to_create = [
            Indicator(
                code=r["code"],
                name=r["name"],
                component=r["component"],
                display_order=r["display_order"],
            )
            for r in summary_rows
            if r["code"] not in existing
        ]
        # Deduplicate by code (keep first)
        seen = set()
        deduped = []
        for o in to_create:
            if o.code not in seen:
                seen.add(o.code)
                deduped.append(o)
        if deduped:
            new_codes = [o.code for o in deduped]
            Indicator.objects.bulk_create(deduped)
            for ind in Indicator.objects.filter(code__in=new_codes).only("id", "code"):
                existing[ind.code] = ind.id

        # Replace IndicatorData for this fiscal_year
        deleted, _ = IndicatorData.objects.filter(fiscal_year=fiscal_year).delete()
        self.stdout.write(f"  Indicator_Summary: replaced {deleted} existing IndicatorData for FY{fiscal_year}")
        self.stdout.write(f"  New indicators: {len(deduped)}")

        indicator_data_list = [
            IndicatorData(
                indicator_id=existing[r["code"]],
                fiscal_year=fiscal_year,
                records_count=_safe_int(r["records"]),
                fallback_records=_safe_int(r["fallback"]),
                gross_lcu_bn=_safe_decimal(r["gross_bn"], Decimal("0")),
                weighted_lcu_bn=_safe_decimal(r["weighted_bn"], Decimal("0")),
                share_weighted_percent=_safe_decimal(r["share_pct"], Decimal("0")),
                status=DataStatus.VALIDATED,
            )
            for r in summary_rows
            if r["code"] in existing
        ]

        for i in range(0, len(indicator_data_list), batch_size):
            IndicatorData.objects.bulk_create(indicator_data_list[i : i + batch_size])
        self.stdout.write(f"  IndicatorData created: {len(indicator_data_list)} rows")

    def _load_mapping_sheet(self, ws, fiscal_year, batch_size):
        """Load Mapping sheet into BudgetLineMapping. Columns A–X."""
        # A=YEAR, B=CODE, C=MDA, D=SUB PROGRAM NAME, E=PROJECT NAME, F=BUDGET LINE, G=TYPE, H=SOURCE,
        # I=SPECIFIC/SUPPORTIVE, J=GROUP, K=Food_system_component, L=Amount_gross_LCU, M=Amount_weighted_LCU,
        # N=Primary_indicator, O=Direct_effect_pathway, P=Key_reference(s), Q=Match_type, R=Notes,
        # S=Indicator_component, T=Indicator, U=Specification, V=Benchmark, W=Gap, X=Responsiveness
        batch = []
        total = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            # Use fiscal_year from argument (file may have mixed years; we store one)
            year = fiscal_year
            if row[0]:
                s = str(row[0]).strip()
                m = re.match(r"^(\d{4})", s)
                if m:
                    year = int(m.group(1))

            batch.append(
                BudgetLineMapping(
                    fiscal_year=year,
                    code=_safe_str(row[1] if len(row) > 1 else None, 50),
                    mda=_safe_str(row[2] if len(row) > 2 else None, 255),
                    sub_program_name=_safe_str(row[3] if len(row) > 3 else None, 255),
                    project_name=_safe_str(row[4] if len(row) > 4 else None, 500),
                    budget_line=_safe_str(row[5] if len(row) > 5 else None, 500),
                    type=_safe_str(row[6] if len(row) > 6 else None, 50),
                    source=_safe_str(row[7] if len(row) > 7 else None, 100),
                    specific_supportive=_safe_str(row[8] if len(row) > 8 else None, 100),
                    group=_safe_str(row[9] if len(row) > 9 else None, 100),
                    food_system_component=_safe_str(row[10] if len(row) > 10 else None, 100),
                    amount_gross_lcu=_safe_decimal(row[11] if len(row) > 11 else None),
                    amount_weighted_lcu=_safe_decimal(row[12] if len(row) > 12 else None),
                    primary_indicator=_safe_str(row[13] if len(row) > 13 else None, 100),
                    direct_effect_pathway=_safe_str(row[14] if len(row) > 14 else None),
                    key_references=_safe_str(row[15] if len(row) > 15 else None),
                    match_type=_safe_str(row[16] if len(row) > 16 else None, 50),
                    notes=_safe_str(row[17] if len(row) > 17 else None),
                    indicator_component=_safe_str(row[18] if len(row) > 18 else None, 50),
                    indicator=_safe_str(row[19] if len(row) > 19 else None, 255),
                    specification=_safe_str(row[20] if len(row) > 20 else None),
                    benchmark=_safe_str(row[21] if len(row) > 21 else None, 100),
                    gap=_safe_str(row[22] if len(row) > 22 else None, 100),
                    responsiveness=_safe_str(row[23] if len(row) > 23 else None, 100),
                )
            )
            if len(batch) >= batch_size:
                BudgetLineMapping.objects.bulk_create(batch)
                total += len(batch)
                batch = []

        if batch:
            BudgetLineMapping.objects.bulk_create(batch)
            total += len(batch)
        self.stdout.write(f"  Mapping sheet: {total} BudgetLineMapping rows created")

    def _load_mapping_per_year(self, ws, single_fiscal_year, batch_size):
        """Aggregate budget data per indicator per year from Mapping sheet and create/update IndicatorData.

        If single_fiscal_year is set, only process that year. Otherwise process all years found.
        Also loads raw BudgetLineMapping rows.
        """
        from collections import defaultdict

        # First pass: aggregate budget data per (year, indicator_code)
        # Structure: {year: {code: {gross, weighted, records, fallback, component}}}
        year_data = defaultdict(lambda: defaultdict(lambda: {
            "gross": Decimal("0"), "weighted": Decimal("0"),
            "records": 0, "fallback": 0, "component": "",
        }))

        mapping_batch = []
        mapping_total = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue

            # Parse year
            s = str(row[0]).strip()
            m = re.match(r"^(\d{4})", s)
            if not m:
                continue
            fy = int(m.group(1))

            if single_fiscal_year and fy != single_fiscal_year:
                continue

            # Parse indicator code
            code, _ = _parse_primary_indicator(row[13] if len(row) > 13 else None)
            if not code:
                continue

            gross = _safe_decimal(row[11] if len(row) > 11 else None, Decimal("0"))
            weighted = _safe_decimal(row[12] if len(row) > 12 else None, Decimal("0"))
            match_type = _safe_str(row[16] if len(row) > 16 else None, 50)

            entry = year_data[fy][code]
            entry["gross"] += gross / Decimal("1000000000")  # LCU -> billions
            entry["weighted"] += weighted / Decimal("1000000000")
            entry["records"] += 1
            if "fallback" in match_type.lower() or "component" in match_type.lower():
                entry["fallback"] += 1
            if not entry["component"]:
                entry["component"] = _component_label_to_value(
                    row[18] if len(row) > 18 else None
                )

            # Build BudgetLineMapping
            mapping_batch.append(
                BudgetLineMapping(
                    fiscal_year=fy,
                    code=_safe_str(row[1] if len(row) > 1 else None, 50),
                    mda=_safe_str(row[2] if len(row) > 2 else None, 255),
                    sub_program_name=_safe_str(row[3] if len(row) > 3 else None, 255),
                    project_name=_safe_str(row[4] if len(row) > 4 else None, 500),
                    budget_line=_safe_str(row[5] if len(row) > 5 else None, 500),
                    type=_safe_str(row[6] if len(row) > 6 else None, 50),
                    source=_safe_str(row[7] if len(row) > 7 else None, 100),
                    specific_supportive=_safe_str(row[8] if len(row) > 8 else None, 100),
                    group=_safe_str(row[9] if len(row) > 9 else None, 100),
                    food_system_component=_safe_str(row[10] if len(row) > 10 else None, 100),
                    amount_gross_lcu=_safe_decimal(row[11] if len(row) > 11 else None),
                    amount_weighted_lcu=_safe_decimal(row[12] if len(row) > 12 else None),
                    primary_indicator=_safe_str(row[13] if len(row) > 13 else None, 100),
                    direct_effect_pathway=_safe_str(row[14] if len(row) > 14 else None),
                    key_references=_safe_str(row[15] if len(row) > 15 else None),
                    match_type=match_type,
                    notes=_safe_str(row[17] if len(row) > 17 else None),
                    indicator_component=_safe_str(row[18] if len(row) > 18 else None, 50),
                    indicator=_safe_str(row[19] if len(row) > 19 else None, 255),
                    specification=_safe_str(row[20] if len(row) > 20 else None),
                    benchmark=_safe_str(row[21] if len(row) > 21 else None, 100),
                    gap=_safe_str(row[22] if len(row) > 22 else None, 100),
                    responsiveness=_safe_str(row[23] if len(row) > 23 else None, 100),
                )
            )
            if len(mapping_batch) >= batch_size:
                BudgetLineMapping.objects.bulk_create(mapping_batch)
                mapping_total += len(mapping_batch)
                mapping_batch = []

        if mapping_batch:
            BudgetLineMapping.objects.bulk_create(mapping_batch)
            mapping_total += len(mapping_batch)

        self.stdout.write(f"  Mapping sheet: {mapping_total} BudgetLineMapping rows created")

        # Now create/update IndicatorData per year
        indicator_lookup = {ind.code: ind for ind in Indicator.objects.all()}
        total_created = 0
        total_updated = 0

        for fy in sorted(year_data.keys()):
            indicators_in_year = year_data[fy]

            # Compute share_weighted_percent per year
            total_weighted = sum(v["weighted"] for v in indicators_in_year.values())

            for code, agg in indicators_in_year.items():
                if code not in indicator_lookup:
                    continue

                share_pct = (
                    (agg["weighted"] / total_weighted * 100)
                    if total_weighted > 0 else Decimal("0")
                )

                # Update existing or create new
                existing = IndicatorData.objects.filter(
                    indicator=indicator_lookup[code], fiscal_year=fy
                ).first()

                if existing:
                    existing.gross_lcu_bn = agg["gross"]
                    existing.weighted_lcu_bn = agg["weighted"]
                    existing.share_weighted_percent = share_pct
                    existing.records_count = agg["records"]
                    existing.fallback_records = agg["fallback"]
                    existing.save(update_fields=[
                        "gross_lcu_bn", "weighted_lcu_bn", "share_weighted_percent",
                        "records_count", "fallback_records",
                    ])
                    total_updated += 1
                else:
                    IndicatorData.objects.create(
                        indicator=indicator_lookup[code],
                        fiscal_year=fy,
                        gross_lcu_bn=agg["gross"],
                        weighted_lcu_bn=agg["weighted"],
                        share_weighted_percent=share_pct,
                        records_count=agg["records"],
                        fallback_records=agg["fallback"],
                        status=DataStatus.VALIDATED,
                    )
                    total_created += 1

            n = len(indicators_in_year)
            self.stdout.write(f"  FY{fy}: {n} indicators (budget from Mapping sheet)")

        self.stdout.write(f"  Budget IndicatorData: {total_created} created, {total_updated} updated")
