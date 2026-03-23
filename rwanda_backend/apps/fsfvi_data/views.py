"""Views for FSFVI indicator data entry."""

import csv
import io
import logging
import re
from decimal import Decimal, InvalidOperation

from django.db import transaction
from django.db.models import Sum, Count
from rest_framework import status
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from .models import Indicator, IndicatorData, IndicatorComponent, DataStatus
from .serializers import (
    IndicatorSerializer,
    IndicatorDataSerializer,
    IndicatorDataInputSerializer,
    BulkIndicatorDataInputSerializer,
    FiscalYearSummarySerializer,
)

logger = logging.getLogger(__name__)

# Regex for parsing indicator codes like "IND-16: Share of production marketed (%)"
INDICATOR_CODE_RE = re.compile(r"^(IND-\d+)\s*:?\s*(.*)$", re.I)


class IndicatorListView(APIView):
    """
    List all indicator definitions.

    GET /api/indicators/
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        component = request.query_params.get("component")
        queryset = Indicator.objects.filter(is_active=True)

        if component:
            queryset = queryset.filter(component=component)

        queryset = queryset.order_by("component", "display_order", "code")
        serializer = IndicatorSerializer(queryset, many=True)
        return Response(serializer.data)


class IndicatorDataListView(APIView):
    """
    List/create indicator data for a fiscal year.

    GET  /api/indicators/data/?fiscal_year=2024
    POST /api/indicators/data/  (single indicator)
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        fiscal_year = request.query_params.get("fiscal_year")
        component = request.query_params.get("component")

        if not fiscal_year:
            return Response(
                {"error": "fiscal_year query parameter is required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            fiscal_year = int(fiscal_year)
        except ValueError:
            return Response(
                {"error": "fiscal_year must be an integer"},
                status=status.HTTP_400_BAD_REQUEST
            )

        queryset = IndicatorData.objects.filter(fiscal_year=fiscal_year)
        if component:
            queryset = queryset.filter(indicator__component=component)

        queryset = queryset.select_related("indicator", "created_by").order_by(
            "indicator__component", "indicator__display_order", "indicator__code"
        )

        serializer = IndicatorDataSerializer(queryset, many=True)
        return Response(serializer.data)

    def post(self, request):
        serializer = IndicatorDataInputSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        data = serializer.validated_data

        try:
            indicator = Indicator.objects.get(id=data["indicator_id"])
        except Indicator.DoesNotExist:
            return Response(
                {"error": f"Indicator {data['indicator_id']} not found"},
                status=status.HTTP_404_NOT_FOUND
            )

        # Create or update
        indicator_data, created = IndicatorData.objects.update_or_create(
            indicator=indicator,
            fiscal_year=data["fiscal_year"],
            defaults={
                "records_count": data.get("records_count", 0),
                "gross_lcu_bn": data["gross_lcu_bn"],
                "weighted_lcu_bn": data["weighted_lcu_bn"],
                "share_weighted_percent": data.get("share_weighted_percent", Decimal("0")),
                "observed_value": data.get("observed_value"),
                "benchmark_value": data.get("benchmark_value"),
                "benchmark_used_type": data.get("benchmark_used_type", ""),
                "financial_allocation_usd": data.get("financial_allocation_usd"),
                "sensitivity_parameter": data.get("sensitivity_parameter"),
                "status": DataStatus.DRAFT,
                "created_by": request.user if created else None,
            }
        )

        result_serializer = IndicatorDataSerializer(indicator_data)
        return Response(
            result_serializer.data,
            status=status.HTTP_201_CREATED if created else status.HTTP_200_OK
        )


class BulkIndicatorDataView(APIView):
    """
    Bulk create/update indicator data for a fiscal year.

    POST /api/indicators/data/bulk/

    Body: {
        "fiscal_year": 2025,
        "indicators": [
            {
                "indicator_id": "uuid",
                "gross_lcu_bn": 100.5,
                "weighted_lcu_bn": 80.2,
                "observed_value": 45.5,
                "benchmark_value": 60.0
            },
            ...
        ]
    }
    """
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request):
        serializer = BulkIndicatorDataInputSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        data = serializer.validated_data
        fiscal_year = data["fiscal_year"]
        indicators_data = data["indicators"]

        # Validate all indicators exist
        indicator_ids = [ind.get("indicator_id") for ind in indicators_data if ind.get("indicator_id")]
        existing_indicators = {
            str(ind.id): ind
            for ind in Indicator.objects.filter(id__in=indicator_ids)
        }

        created_count = 0
        updated_count = 0
        errors = []
        results = []

        for i, ind_data in enumerate(indicators_data):
            ind_id = str(ind_data.get("indicator_id", ""))

            if ind_id not in existing_indicators:
                errors.append({"index": i, "error": f"Indicator {ind_id} not found"})
                continue

            indicator = existing_indicators[ind_id]

            try:
                gross_lcu = Decimal(str(ind_data.get("gross_lcu_bn", 0)))
                weighted_lcu = Decimal(str(ind_data.get("weighted_lcu_bn", 0)))
                observed = ind_data.get("observed_value")
                benchmark = ind_data.get("benchmark_value")

                observed_val = Decimal(str(observed)) if observed is not None else None
                benchmark_val = Decimal(str(benchmark)) if benchmark is not None else None

            except (ValueError, TypeError) as e:
                errors.append({"index": i, "indicator_id": ind_id, "error": str(e)})
                continue

            indicator_data, created = IndicatorData.objects.update_or_create(
                indicator=indicator,
                fiscal_year=fiscal_year,
                defaults={
                    "records_count": ind_data.get("records_count", 0),
                    "gross_lcu_bn": gross_lcu,
                    "weighted_lcu_bn": weighted_lcu,
                    "share_weighted_percent": Decimal(str(ind_data.get("share_weighted_percent", 0))),
                    "observed_value": observed_val,
                    "benchmark_value": benchmark_val,
                    "benchmark_used_type": ind_data.get("benchmark_used_type", ""),
                    "financial_allocation_usd": ind_data.get("financial_allocation_usd"),
                    "sensitivity_parameter": ind_data.get("sensitivity_parameter"),
                    "status": DataStatus.DRAFT,
                }
            )

            if created:
                indicator_data.created_by = request.user
                indicator_data.save(update_fields=["created_by"])
                created_count += 1
            else:
                updated_count += 1

            results.append(IndicatorDataSerializer(indicator_data).data)

        # Recalculate share_weighted_percent based on totals
        total_weighted = IndicatorData.objects.filter(
            fiscal_year=fiscal_year
        ).aggregate(total=Sum("weighted_lcu_bn"))["total"] or Decimal("0")

        if total_weighted > 0:
            for ind_data in IndicatorData.objects.filter(fiscal_year=fiscal_year):
                ind_data.share_weighted_percent = (ind_data.weighted_lcu_bn / total_weighted) * 100
                ind_data.save(update_fields=["share_weighted_percent"])

        return Response({
            "fiscal_year": fiscal_year,
            "created": created_count,
            "updated": updated_count,
            "errors": errors,
            "total_processed": len(results),
        }, status=status.HTTP_200_OK if not errors else status.HTTP_207_MULTI_STATUS)


class FiscalYearSummaryView(APIView):
    """
    Get summary of indicator data for a fiscal year.

    GET /api/indicators/data/summary/?fiscal_year=2024
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        fiscal_year = request.query_params.get("fiscal_year")

        if not fiscal_year:
            return Response(
                {"error": "fiscal_year query parameter is required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            fiscal_year = int(fiscal_year)
        except ValueError:
            return Response(
                {"error": "fiscal_year must be an integer"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Get total indicators
        total_indicators = Indicator.objects.filter(is_active=True).count()

        # Get indicator data stats
        data_queryset = IndicatorData.objects.filter(fiscal_year=fiscal_year)
        indicators_with_data = data_queryset.count()

        totals = data_queryset.aggregate(
            total_gross=Sum("gross_lcu_bn"),
            total_weighted=Sum("weighted_lcu_bn"),
        )

        # Status counts
        status_counts = dict(
            data_queryset.values("status").annotate(count=Count("id")).values_list("status", "count")
        )

        # Component summary
        components_summary = []
        for comp_value, comp_display in IndicatorComponent.choices:
            comp_data = data_queryset.filter(indicator__component=comp_value)
            comp_agg = comp_data.aggregate(
                count=Count("id"),
                gross=Sum("gross_lcu_bn"),
                weighted=Sum("weighted_lcu_bn"),
            )
            comp_total_indicators = Indicator.objects.filter(
                is_active=True, component=comp_value
            ).count()

            components_summary.append({
                "component": comp_value,
                "component_display": comp_display,
                "total_indicators": comp_total_indicators,
                "indicators_with_data": comp_agg["count"] or 0,
                "gross_lcu_bn": float(comp_agg["gross"] or 0),
                "weighted_lcu_bn": float(comp_agg["weighted"] or 0),
            })

        return Response({
            "fiscal_year": fiscal_year,
            "total_indicators": total_indicators,
            "indicators_with_data": indicators_with_data,
            "total_gross_lcu_bn": float(totals["total_gross"] or 0),
            "total_weighted_lcu_bn": float(totals["total_weighted"] or 0),
            "status_counts": status_counts,
            "components_summary": components_summary,
        })


class AvailableDataYearsView(APIView):
    """
    List fiscal years that have indicator data.

    GET /api/indicators/data/available-years/
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        years = (
            IndicatorData.objects
            .values_list("fiscal_year", flat=True)
            .distinct()
            .order_by("-fiscal_year")
        )
        return Response({"fiscal_years": list(years)})


class CopyFiscalYearDataView(APIView):
    """
    Copy indicator data from one fiscal year to another (as starting point).

    POST /api/indicators/data/copy/

    Body: {
        "source_fiscal_year": 2024,
        "target_fiscal_year": 2025
    }
    """
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request):
        source_year = request.data.get("source_fiscal_year")
        target_year = request.data.get("target_fiscal_year")

        if not source_year or not target_year:
            return Response(
                {"error": "source_fiscal_year and target_fiscal_year are required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            source_year = int(source_year)
            target_year = int(target_year)
        except ValueError:
            return Response(
                {"error": "Fiscal years must be integers"},
                status=status.HTTP_400_BAD_REQUEST
            )

        if source_year == target_year:
            return Response(
                {"error": "Source and target fiscal years must be different"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Check if target year already has data
        existing_count = IndicatorData.objects.filter(fiscal_year=target_year).count()
        if existing_count > 0:
            return Response(
                {"error": f"Target fiscal year {target_year} already has {existing_count} indicator records. Delete them first or update individually."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Copy data
        source_data = IndicatorData.objects.filter(fiscal_year=source_year)
        if not source_data.exists():
            return Response(
                {"error": f"No indicator data found for source fiscal year {source_year}"},
                status=status.HTTP_404_NOT_FOUND
            )

        created_count = 0
        for src in source_data:
            IndicatorData.objects.create(
                indicator=src.indicator,
                fiscal_year=target_year,
                records_count=src.records_count,
                fallback_records=src.fallback_records,
                gross_lcu_bn=src.gross_lcu_bn,
                weighted_lcu_bn=src.weighted_lcu_bn,
                share_weighted_percent=src.share_weighted_percent,
                observed_value=src.observed_value,
                benchmark_value=src.benchmark_value,
                benchmark_used_type=src.benchmark_used_type,
                financial_allocation_usd=src.financial_allocation_usd,
                sensitivity_parameter=src.sensitivity_parameter,
                status=DataStatus.DRAFT,
                created_by=request.user,
            )
            created_count += 1

        return Response({
            "source_fiscal_year": source_year,
            "target_fiscal_year": target_year,
            "copied_records": created_count,
        }, status=status.HTTP_201_CREATED)


class DeleteFiscalYearDataView(APIView):
    """
    Delete all indicator data for a fiscal year.

    DELETE /api/indicators/data/delete-year/?fiscal_year=2025
    """
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def delete(self, request):
        fiscal_year = request.query_params.get("fiscal_year")

        if not fiscal_year:
            return Response(
                {"error": "fiscal_year query parameter is required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            fiscal_year = int(fiscal_year)
        except ValueError:
            return Response(
                {"error": "fiscal_year must be an integer"},
                status=status.HTTP_400_BAD_REQUEST
            )

        deleted_count, _ = IndicatorData.objects.filter(fiscal_year=fiscal_year).delete()

        return Response({
            "fiscal_year": fiscal_year,
            "deleted_records": deleted_count,
        })


def _safe_decimal(value, default=None):
    """Convert value to Decimal safely."""
    if value is None or value == "":
        return default
    if isinstance(value, Decimal):
        return value
    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, TypeError, ValueError):
        return default


def _parse_indicator_code(value):
    """Return (code, name) e.g. ('IND-16', 'Share of production marketed (%)')."""
    if not value or not str(value).strip():
        return None, None
    s = str(value).strip()
    m = INDICATOR_CODE_RE.match(s)
    if m:
        return m.group(1).strip().upper(), (m.group(2) or "").strip() or s
    # Check if it's just a code like "IND-16"
    if s.upper().startswith("IND-"):
        return s.split(":")[0].strip().upper(), s
    return None, s


def _normalize_column_name(name):
    """Normalize column name for matching."""
    if not name:
        return ""
    return str(name).strip().lower().replace(" ", "_").replace("-", "_")


class BulkFileImportView(APIView):
    """
    Import indicator data from CSV or Excel file.

    POST /api/indicators/data/import/

    Supports two formats:
    1. Simple CSV/Excel with columns:
       - indicator_code (required): e.g., "IND-01" or "IND-01: Yield (t/ha)"
       - gross_lcu_bn (required): Gross budget in billions LCU
       - weighted_lcu_bn (required): Weighted budget in billions LCU
       - observed_value (optional): Current performance value
       - benchmark_value (optional): Target/benchmark value

    2. Full Indicator_Summary format (Excel):
       - Primary_indicator, Indicator_component, Indicator, Records, Fallback_records,
         Gross_LCU_bn, Weighted_LCU_bn, Share_weighted_%

    Form data:
    - file: The CSV or Excel file
    - fiscal_year: Target fiscal year (required)
    - mode: "preview" (default) or "import"
    """
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        uploaded_file = request.FILES.get("file")
        fiscal_year = request.data.get("fiscal_year")
        mode = request.data.get("mode", "preview")

        if not uploaded_file:
            return Response(
                {"error": "No file uploaded"},
                status=status.HTTP_400_BAD_REQUEST
            )

        if not fiscal_year:
            return Response(
                {"error": "fiscal_year is required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            fiscal_year = int(fiscal_year)
        except ValueError:
            return Response(
                {"error": "fiscal_year must be an integer"},
                status=status.HTTP_400_BAD_REQUEST
            )

        filename = uploaded_file.name.lower()

        try:
            if filename.endswith(".csv"):
                rows = self._parse_csv(uploaded_file)
            elif filename.endswith((".xlsx", ".xls")):
                rows = self._parse_excel(uploaded_file)
            else:
                return Response(
                    {"error": "Unsupported file format. Use CSV or Excel (.xlsx, .xls)"},
                    status=status.HTTP_400_BAD_REQUEST
                )
        except Exception as e:
            logger.exception("Failed to parse uploaded file")
            return Response(
                {"error": f"Failed to parse file: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST
            )

        if not rows:
            return Response(
                {"error": "No data rows found in file"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Process rows and match to indicators
        result = self._process_rows(rows, fiscal_year, mode, request.user)

        if mode == "import" and not result.get("errors"):
            return Response(result, status=status.HTTP_201_CREATED)
        elif result.get("errors"):
            return Response(result, status=status.HTTP_207_MULTI_STATUS)
        else:
            return Response(result, status=status.HTTP_200_OK)

    def _parse_csv(self, uploaded_file):
        """Parse CSV file and return list of row dicts."""
        content = uploaded_file.read().decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(content))

        # Normalize column names
        rows = []
        for row in reader:
            normalized_row = {
                _normalize_column_name(k): v
                for k, v in row.items()
            }
            rows.append(normalized_row)

        return rows

    def _parse_excel(self, uploaded_file):
        """Parse Excel file and return list of row dicts."""
        try:
            import openpyxl
        except ImportError:
            raise Exception("openpyxl is required for Excel import. Install with: pip install openpyxl")

        wb = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)

        # Try to find the right sheet
        sheet_name = None
        for name in ["Indicator_Summary", "Data", "Sheet1", wb.sheetnames[0]]:
            if name in wb.sheetnames:
                sheet_name = name
                break

        ws = wb[sheet_name]

        # Get headers from first row
        headers = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True)):
            headers.append(_normalize_column_name(cell) if cell else "")

        # Parse data rows
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not any(row):
                continue
            row_dict = {}
            for i, value in enumerate(row):
                if i < len(headers) and headers[i]:
                    row_dict[headers[i]] = value
            rows.append(row_dict)

        wb.close()
        return rows

    def _process_rows(self, rows, fiscal_year, mode, user):
        """Process parsed rows and optionally import them."""
        # Build indicator lookup by code
        indicator_lookup = {
            ind.code.upper(): ind
            for ind in Indicator.objects.filter(is_active=True)
        }

        # Column name mappings (support various formats)
        code_columns = ["indicator_code", "primary_indicator", "code", "ind_code"]
        gross_columns = ["gross_lcu_bn", "gross", "gross_bn", "amount_gross"]
        weighted_columns = ["weighted_lcu_bn", "weighted", "weighted_bn", "amount_weighted"]
        observed_columns = ["observed_value", "observed", "current_value", "current"]
        benchmark_columns = ["benchmark_value", "benchmark", "target_value", "target"]

        def find_column(row, candidates):
            for col in candidates:
                if col in row and row[col] is not None and str(row[col]).strip():
                    return row[col]
            return None

        preview_data = []
        errors = []
        matched_count = 0
        unmatched_count = 0

        for i, row in enumerate(rows):
            row_num = i + 2  # Account for header row

            # Find indicator code
            code_value = find_column(row, code_columns)
            if not code_value:
                errors.append({
                    "row": row_num,
                    "error": "Missing indicator code"
                })
                continue

            code, name = _parse_indicator_code(code_value)
            if not code:
                errors.append({
                    "row": row_num,
                    "error": f"Invalid indicator code format: {code_value}"
                })
                unmatched_count += 1
                continue

            if code not in indicator_lookup:
                errors.append({
                    "row": row_num,
                    "error": f"Indicator not found: {code}"
                })
                unmatched_count += 1
                continue

            indicator = indicator_lookup[code]

            # Parse values
            gross = _safe_decimal(find_column(row, gross_columns), Decimal("0"))
            weighted = _safe_decimal(find_column(row, weighted_columns), Decimal("0"))
            observed = _safe_decimal(find_column(row, observed_columns))
            benchmark = _safe_decimal(find_column(row, benchmark_columns))

            matched_count += 1

            preview_data.append({
                "row": row_num,
                "indicator_code": code,
                "indicator_name": indicator.name,
                "component": indicator.component,
                "gross_lcu_bn": float(gross),
                "weighted_lcu_bn": float(weighted),
                "observed_value": float(observed) if observed is not None else None,
                "benchmark_value": float(benchmark) if benchmark is not None else None,
            })

        result = {
            "fiscal_year": fiscal_year,
            "mode": mode,
            "total_rows": len(rows),
            "matched": matched_count,
            "unmatched": unmatched_count,
            "errors": errors,
            "preview": preview_data[:100] if mode == "preview" else [],  # Limit preview
        }

        if mode == "import" and matched_count > 0:
            with transaction.atomic():
                created_count = 0
                updated_count = 0

                for item in preview_data:
                    indicator = indicator_lookup[item["indicator_code"]]

                    observed_val = (
                        Decimal(str(item["observed_value"]))
                        if item["observed_value"] is not None else None
                    )
                    benchmark_val = (
                        Decimal(str(item["benchmark_value"]))
                        if item["benchmark_value"] is not None else None
                    )

                    _, created = IndicatorData.objects.update_or_create(
                        indicator=indicator,
                        fiscal_year=fiscal_year,
                        defaults={
                            "gross_lcu_bn": Decimal(str(item["gross_lcu_bn"])),
                            "weighted_lcu_bn": Decimal(str(item["weighted_lcu_bn"])),
                            "observed_value": observed_val,
                            "benchmark_value": benchmark_val,
                            "status": DataStatus.DRAFT,
                            "created_by": user if created else None,
                        }
                    )

                    if created:
                        created_count += 1
                    else:
                        updated_count += 1

                # Recalculate share percentages
                total_weighted = IndicatorData.objects.filter(
                    fiscal_year=fiscal_year
                ).aggregate(total=Sum("weighted_lcu_bn"))["total"] or Decimal("0")

                if total_weighted > 0:
                    for ind_data in IndicatorData.objects.filter(fiscal_year=fiscal_year):
                        ind_data.share_weighted_percent = (
                            ind_data.weighted_lcu_bn / total_weighted * 100
                        )
                        ind_data.save(update_fields=["share_weighted_percent"])

                result["created"] = created_count
                result["updated"] = updated_count

        return result


class DownloadTemplateView(APIView):
    """
    Download a CSV template for bulk import.

    GET /api/indicators/data/template/
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from django.http import HttpResponse

        # Create CSV content
        output = io.StringIO()
        writer = csv.writer(output)

        # Write header
        writer.writerow([
            "indicator_code",
            "gross_lcu_bn",
            "weighted_lcu_bn",
            "observed_value",
            "benchmark_value"
        ])

        # Write indicator rows with empty data
        indicators = Indicator.objects.filter(is_active=True).order_by(
            "component", "display_order", "code"
        )

        for ind in indicators:
            writer.writerow([
                f"{ind.code}: {ind.name}",
                "",  # gross_lcu_bn
                "",  # weighted_lcu_bn
                "",  # observed_value
                "",  # benchmark_value
            ])

        response = HttpResponse(
            output.getvalue(),
            content_type="text/csv",
            headers={"Content-Disposition": 'attachment; filename="indicator_data_template.csv"'},
        )
        return response
